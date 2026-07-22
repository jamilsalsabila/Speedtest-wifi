from __future__ import annotations

import argparse
import csv
import json
import logging
import mimetypes
import os
import platform
import re
import shutil
import smtplib
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

BASE_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
CONFIG_FILE = BASE_DIR / "config.json"
DATA_DIR = BASE_DIR / "data"
REPORT_DIR = BASE_DIR / "reports"
LOG_DIR = BASE_DIR / "logs"
CSV_FILE = DATA_DIR / "speedtest_log.csv"
_STANDARD_STREAM_REFS: list[Any] = []

HEADERS = [
    "tanggal",
    "waktu",
    "komputer",
    "wifi",
    "download_mbps",
    "upload_mbps",
    "ping_ms",
    "status",
    "error_type",
    "error",
]

ERROR_SSID_NOT_FOUND = "SSID_NOT_FOUND"
ERROR_AUTH_FAILED = "AUTH_FAILED"
ERROR_WIFI_ADAPTER = "WIFI_ADAPTER_ERROR"
ERROR_CONNECT_FAILED = "CONNECT_FAILED"
ERROR_NOT_CONNECTED = "NOT_CONNECTED"
ERROR_NO_INTERNET = "NO_INTERNET"
ERROR_SPEEDTEST_FAILED = "SPEEDTEST_FAILED"
ERROR_REPORT_FAILED = "REPORT_FAILED"
ERROR_EMAIL_FAILED = "EMAIL_FAILED"
ERROR_UNKNOWN = "UNKNOWN"


class WifiMonitorError(RuntimeError):
    def __init__(self, error_type: str, message: str) -> None:
        super().__init__(message)
        self.error_type = error_type


@dataclass(frozen=True)
class WifiProfile:
    ssid: str
    password: str = ""
    label: str = ""

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "WifiProfile":
        ssid = str(data.get("ssid", "")).strip()
        if not ssid:
            raise ValueError("Setiap Wi-Fi harus memiliki ssid.")
        return cls(
            ssid=ssid,
            password=str(data.get("password", "")),
            label=str(data.get("label") or ssid),
        )


def setup() -> None:
    for folder in (DATA_DIR, REPORT_DIR, LOG_DIR):
        folder.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        filename=LOG_DIR / "monitor.log",
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        encoding="utf-8",
    )

    if CSV_FILE.exists():
        migrate_csv_headers()
    else:
        with CSV_FILE.open("w", newline="", encoding="utf-8-sig") as f:
            csv.DictWriter(f, fieldnames=HEADERS).writeheader()


def migrate_csv_headers() -> None:
    with CSV_FILE.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        old_headers = reader.fieldnames or []

    if old_headers == HEADERS:
        return

    normalized = []
    for row in rows:
        normalized.append({header: row.get(header, "") for header in HEADERS})

    with CSV_FILE.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(normalized)


def run(
    args: list[str],
    timeout: int = 240,
    env: dict[str, str] | None = None,
) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        args,
        capture_output=True,
        text=True,
        timeout=timeout,
        check=False,
        env=env,
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )


def command_output(result: subprocess.CompletedProcess[str]) -> str:
    return (result.stderr or "").strip() or (result.stdout or "").strip()


def command_stdout_lines(result: subprocess.CompletedProcess[str]) -> list[str]:
    return (result.stdout or "").splitlines()


def load_config(config_file: Path = CONFIG_FILE) -> dict[str, Any]:
    if not config_file.exists():
        raise FileNotFoundError(
            f"{config_file.name} belum dibuat. Jalankan GUI atau salin config.example.json."
        )

    with config_file.open(encoding="utf-8") as f:
        config = json.load(f)

    profiles = config.get("wifi_profiles") or []
    config["wifi_profiles"] = [profile.__dict__ for profile in map(WifiProfile.from_dict, profiles)]
    if not config["wifi_profiles"] and not bool(config.get("test_current_connection", False)):
        raise ValueError("Isi minimal satu Wi-Fi atau aktifkan tes Ethernet/koneksi aktif.")

    return config


def command_exists(name: str) -> bool:
    return shutil.which(name) is not None


def connect_wifi(profile: WifiProfile, settle_seconds: int, retries: int = 2) -> None:
    system = platform.system().lower()
    logging.info("Menghubungkan ke SSID %s di %s", profile.ssid, system)

    available = available_wifi_ssids()
    if available is not None and profile.ssid not in available:
        raise WifiMonitorError(ERROR_SSID_NOT_FOUND, f"SSID '{profile.ssid}' tidak ditemukan di daftar Wi-Fi sekitar.")

    attempts = max(1, retries + 1)
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            if system == "windows":
                connect_wifi_windows(profile)
            elif system == "darwin":
                connect_wifi_macos(profile)
            elif system == "linux":
                connect_wifi_linux(profile)
            else:
                raise WifiMonitorError(ERROR_UNKNOWN, f"OS belum didukung: {platform.system()}")

            time.sleep(settle_seconds)
            if is_connected_to(profile.ssid):
                return

            raise WifiMonitorError(ERROR_NOT_CONNECTED, f"Komputer belum tersambung ke {profile.ssid}.")
        except Exception as exc:
            last_error = exc
            logging.warning("Percobaan koneksi %s/%s ke %s gagal: %s", attempt, attempts, profile.ssid, exc)
            if attempt < attempts:
                time.sleep(min(10, max(2, settle_seconds // 2)))

    error_type = classify_wifi_error(last_error)
    message = friendly_wifi_error(profile.ssid, error_type, str(last_error or "Koneksi Wi-Fi gagal."))
    raise WifiMonitorError(error_type, message)


def available_wifi_ssids() -> set[str] | None:
    system = platform.system().lower()
    try:
        if system == "windows":
            result = run(["netsh", "wlan", "show", "networks"], timeout=45)
            if result.returncode != 0:
                return None
            ssids = set()
            for line in command_stdout_lines(result):
                if line.strip().lower().startswith("ssid") and ":" in line:
                    ssid = line.split(":", 1)[1].strip()
                    if ssid:
                        ssids.add(ssid)
            return ssids

        if system == "darwin":
            airport = Path("/System/Library/PrivateFrameworks/Apple80211.framework/Versions/Current/Resources/airport")
            if not airport.exists():
                return None
            result = run([str(airport), "-s"], timeout=45)
            if result.returncode != 0:
                return None
            ssids = set()
            for line in command_stdout_lines(result)[1:]:
                text = line.strip()
                if not text:
                    continue
                ssid = text.rsplit(maxsplit=6)[0].strip()
                if ssid:
                    ssids.add(ssid)
            return ssids

        if system == "linux":
            if not command_exists("nmcli"):
                return None
            result = run(["nmcli", "-t", "-f", "SSID", "dev", "wifi", "list"], timeout=45)
            if result.returncode != 0:
                return None
            return {line.replace("\\:", ":").strip() for line in command_stdout_lines(result) if line.strip()}
    except Exception:
        logging.exception("Gagal memindai daftar SSID.")
        return None

    return None


def classify_wifi_error(exc: Exception | None) -> str:
    if isinstance(exc, WifiMonitorError):
        return exc.error_type

    text = str(exc or "").lower()
    auth_markers = [
        "password",
        "authentication",
        "auth",
        "invalid key",
        "incorrect",
        "not authorized",
        "could not be joined",
        "secrets were required",
        "no secrets",
        "802-11-wireless-security",
    ]
    adapter_markers = ["no wireless interface", "wi-fi power", "wifi power", "radio", "adapter", "device not found"]

    if any(marker in text for marker in auth_markers):
        return ERROR_AUTH_FAILED
    if any(marker in text for marker in adapter_markers):
        return ERROR_WIFI_ADAPTER
    if "not connected" in text or "belum tersambung" in text:
        return ERROR_NOT_CONNECTED
    return ERROR_CONNECT_FAILED


def friendly_wifi_error(ssid: str, error_type: str, detail: str) -> str:
    messages = {
        ERROR_SSID_NOT_FOUND: f"SSID '{ssid}' tidak ditemukan. Pastikan Wi-Fi aktif dan SSID berada dalam jangkauan.",
        ERROR_AUTH_FAILED: f"Gagal login ke '{ssid}'. Kemungkinan password salah atau keamanan Wi-Fi tidak cocok.",
        ERROR_WIFI_ADAPTER: "Adapter Wi-Fi bermasalah atau sedang nonaktif.",
        ERROR_NOT_CONNECTED: f"Komputer belum tersambung ke '{ssid}' setelah percobaan koneksi.",
        ERROR_CONNECT_FAILED: f"Gagal menghubungkan ke '{ssid}'.",
    }
    base = messages.get(error_type, f"Gagal menghubungkan ke '{ssid}'.")
    return f"{base} Detail: {detail}".strip()


def connect_wifi_windows(profile: WifiProfile) -> None:
    if profile.password:
        add_windows_profile(profile)

    errors = []
    for args in windows_connect_commands(profile):
        result = run(args, timeout=45)
        if result.returncode == 0:
            return
        output = command_output(result)
        if output:
            errors.append(output)

    raise RuntimeError(errors[-1] if errors else f"Windows gagal menghubungkan ke {profile.ssid}.")


def windows_connect_commands(profile: WifiProfile) -> list[list[str]]:
    profile_names = [profile.ssid]
    for name in windows_profile_names_for_ssid(profile.ssid):
        if name not in profile_names:
            profile_names.append(name)

    interfaces = [None]
    for interface_name in get_windows_wifi_interfaces():
        if interface_name not in interfaces:
            interfaces.append(interface_name)

    commands = []
    for profile_name in profile_names:
        for interface in interfaces:
            command = ["netsh", "wlan", "connect", f"name={profile_name}", f"ssid={profile.ssid}"]
            if interface:
                command.append(f"interface={interface}")
            commands.append(command)

    for profile_name in profile_names:
        for interface in interfaces:
            command = ["netsh", "wlan", "connect", f"name={profile_name}"]
            if interface:
                command.append(f"interface={interface}")
            commands.append(command)

    unique_commands = []
    seen = set()
    for command in commands:
        key = tuple(command)
        if key not in seen:
            unique_commands.append(command)
            seen.add(key)
    return unique_commands


def windows_profile_names_for_ssid(ssid: str) -> list[str]:
    result = run(["netsh", "wlan", "show", "profiles"], timeout=30)
    if result.returncode != 0:
        return []

    names = []
    for line in command_stdout_lines(result):
        text = line.strip()
        lowered = text.lower()
        if ":" not in text or ("profile" not in lowered and "profil" not in lowered):
            continue
        name = text.split(":", 1)[1].strip()
        if not name or name in names:
            continue
        if name_matches_ssid(name, ssid) or windows_profile_has_ssid(name, ssid):
            names.append(name)
    return names


def windows_profile_has_ssid(profile_name: str, ssid: str) -> bool:
    result = run(["netsh", "wlan", "show", "profile", f"name={profile_name}"], timeout=30)
    if result.returncode != 0:
        return False

    for line in command_stdout_lines(result):
        text = line.strip()
        if ":" not in text or "ssid" not in text.lower():
            continue
        value = text.split(":", 1)[1].strip().strip('"')
        if name_matches_ssid(value, ssid):
            return True
    return False


def name_matches_ssid(name: str, ssid: str) -> bool:
    normalized_name = normalize_wifi_name(name)
    normalized_ssid = normalize_wifi_name(ssid)
    return normalized_name == normalized_ssid or normalized_ssid in normalized_name


def normalize_wifi_name(value: str) -> str:
    return re.sub(r"\s+", "", value).lower()


def get_windows_wifi_interface() -> str | None:
    interfaces = get_windows_wifi_interfaces()
    return interfaces[0] if interfaces else None


def get_windows_wifi_interfaces() -> list[str]:
    result = run(["netsh", "wlan", "show", "interfaces"], timeout=30)
    if result.returncode != 0:
        return windows_common_wifi_interfaces()

    interfaces = []
    for line in command_stdout_lines(result):
        text = line.strip()
        label = text.split(":", 1)[0].strip().lower() if ":" in text else ""
        if label in {"name", "nama"}:
            name = text.split(":", 1)[1].strip()
            if name and name not in interfaces:
                interfaces.append(name)

    for name in windows_common_wifi_interfaces():
        if name not in interfaces:
            interfaces.append(name)
    return interfaces


def windows_common_wifi_interfaces() -> list[str]:
    return ["Wi-Fi", "WiFi", "Wireless Network Connection"]


def add_windows_profile(profile: WifiProfile) -> None:
    import tempfile
    import xml.sax.saxutils as xml_utils

    ssid = xml_utils.escape(profile.ssid)
    password = xml_utils.escape(profile.password)
    profile_xml = f"""<?xml version="1.0"?>
<WLANProfile xmlns="http://www.microsoft.com/networking/WLAN/profile/v1">
    <name>{ssid}</name>
    <SSIDConfig><SSID><name>{ssid}</name></SSID></SSIDConfig>
    <connectionType>ESS</connectionType>
    <connectionMode>auto</connectionMode>
    <MSM>
        <security>
            <authEncryption>
                <authentication>WPA2PSK</authentication>
                <encryption>AES</encryption>
                <useOneX>false</useOneX>
            </authEncryption>
            <sharedKey>
                <keyType>passPhrase</keyType>
                <protected>false</protected>
                <keyMaterial>{password}</keyMaterial>
            </sharedKey>
        </security>
    </MSM>
</WLANProfile>
"""
    with tempfile.NamedTemporaryFile("w", suffix=".xml", delete=False, encoding="utf-8") as temp:
        temp.write(profile_xml)
        temp_path = temp.name

    try:
        result = run(["netsh", "wlan", "add", "profile", f"filename={temp_path}", "user=current"], timeout=45)
        if result.returncode != 0:
            output = command_output(result)
            if is_windows_existing_profile_error(output):
                logging.info("Profil Wi-Fi Windows sudah ada untuk %s; memakai profil yang tersedia.", profile.ssid)
                return
            raise RuntimeError(output)
    finally:
        Path(temp_path).unlink(missing_ok=True)


def is_windows_existing_profile_error(output: str) -> bool:
    text = output.lower()
    markers = [
        "already exists",
        "group policy",
        "different user scope",
        "cannot be overwritten",
        "profile already exists",
    ]
    return any(marker in text for marker in markers)


def connect_wifi_macos(profile: WifiProfile) -> None:
    device = get_macos_wifi_device()
    args = ["networksetup", "-setairportnetwork", device, profile.ssid]
    if profile.password:
        args.append(profile.password)

    result = run(args, timeout=60)
    if result.returncode != 0:
        raise RuntimeError(command_output(result))


def get_macos_wifi_device() -> str:
    result = run(["networksetup", "-listallhardwareports"], timeout=30)
    lines = command_stdout_lines(result)
    for index, line in enumerate(lines):
        if "Hardware Port: Wi-Fi" in line or "Hardware Port: AirPort" in line:
            for detail in lines[index + 1 : index + 4]:
                if detail.strip().startswith("Device:"):
                    return detail.split(":", 1)[1].strip()
    raise RuntimeError("Tidak dapat menemukan perangkat Wi-Fi macOS.")


def connect_wifi_linux(profile: WifiProfile) -> None:
    if not command_exists("nmcli"):
        raise RuntimeError("Linux membutuhkan NetworkManager CLI (`nmcli`) untuk koneksi Wi-Fi.")

    args = ["nmcli", "device", "wifi", "connect", profile.ssid]
    if profile.password:
        args.extend(["password", profile.password])

    result = run(args, timeout=60)
    if result.returncode != 0:
        output = command_output(result)
        if is_linux_existing_profile_error(output):
            logging.info("Koneksi Wi-Fi Linux sudah ada untuk %s; memakai koneksi NetworkManager yang tersedia.", profile.ssid)
            if activate_existing_linux_connection(profile):
                return
        raise RuntimeError(output)


def is_linux_existing_profile_error(output: str) -> bool:
    text = output.lower()
    markers = [
        "already exists",
        "connection already exists",
        "exists with uuid",
        "duplicate",
    ]
    return any(marker in text for marker in markers)


def activate_existing_linux_connection(profile: WifiProfile) -> bool:
    connection_name = find_linux_wifi_connection(profile.ssid)
    if not connection_name:
        return False

    if profile.password:
        result = run(
            [
                "nmcli",
                "connection",
                "modify",
                connection_name,
                "wifi-sec.key-mgmt",
                "wpa-psk",
                "wifi-sec.psk",
                profile.password,
            ],
            timeout=45,
        )
        if result.returncode != 0:
            raise RuntimeError(command_output(result))

    result = run(["nmcli", "connection", "up", connection_name], timeout=60)
    if result.returncode != 0:
        raise RuntimeError(command_output(result))
    return True


def find_linux_wifi_connection(ssid: str) -> str | None:
    result = run(
        ["nmcli", "-t", "-f", "NAME,TYPE,802-11-wireless.ssid", "connection", "show"],
        timeout=30,
    )
    if result.returncode != 0:
        return None

    for line in command_stdout_lines(result):
        fields = parse_nmcli_terse_line(line)
        if len(fields) < 2 or fields[1] != "802-11-wireless":
            continue

        name = fields[0]
        saved_ssid = fields[2] if len(fields) > 2 else ""
        if saved_ssid == ssid or name == ssid:
            return name
    return None


def parse_nmcli_terse_line(line: str) -> list[str]:
    fields = []
    current = []
    escaped = False

    for char in line:
        if escaped:
            current.append(char)
            escaped = False
        elif char == "\\":
            escaped = True
        elif char == ":":
            fields.append("".join(current))
            current = []
        else:
            current.append(char)

    if escaped:
        current.append("\\")
    fields.append("".join(current))
    return fields


def current_wifi_ssid() -> str | None:
    system = platform.system().lower()

    if system == "windows":
        result = run(["netsh", "wlan", "show", "interfaces"], timeout=30)
        if result.returncode != 0:
            return None
        for line in command_stdout_lines(result):
            text = line.strip()
            if text.lower().startswith("ssid") and "bssid" not in text.lower() and ":" in text:
                ssid = text.split(":", 1)[1].strip()
                return ssid or None
        return None

    if system == "darwin":
        device = get_macos_wifi_device()
        result = run(["networksetup", "-getairportnetwork", device], timeout=30)
        stdout = result.stdout or ""
        if result.returncode != 0 or "not associated" in stdout.lower():
            return None
        if ":" in stdout:
            ssid = stdout.split(":", 1)[1].strip()
            return ssid or None
        return None

    if system == "linux":
        if not command_exists("nmcli"):
            return None
        result = run(["nmcli", "-t", "-f", "active,ssid", "dev", "wifi"], timeout=30)
        if result.returncode != 0:
            return None
        for line in command_stdout_lines(result):
            if line.lower().startswith("yes:"):
                return line.split(":", 1)[1].replace("\\:", ":").strip() or None

    return None


def disconnect_wifi() -> None:
    system = platform.system().lower()

    if system == "windows":
        interface_name = get_windows_wifi_interface()
        commands = []
        if interface_name:
            commands.append(["netsh", "wlan", "disconnect", f"interface={interface_name}"])
        commands.append(["netsh", "wlan", "disconnect", "interface=*"])
        commands.append(["netsh", "wlan", "disconnect"])

        result = None
        for command in commands:
            result = run(command, timeout=30)
            if result.returncode == 0:
                break
    elif system == "darwin":
        result = run(["networksetup", "-setairportpower", get_macos_wifi_device(), "off"], timeout=30)
        if result.returncode == 0:
            run(["networksetup", "-setairportpower", get_macos_wifi_device(), "on"], timeout=30)
    elif system == "linux":
        device = get_linux_wifi_device()
        if device is None:
            return
        result = run(["nmcli", "device", "disconnect", device], timeout=30)
    else:
        return

    if result.returncode != 0:
        raise RuntimeError(command_output(result))


def get_linux_wifi_device() -> str | None:
    if not command_exists("nmcli"):
        return None
    result = run(["nmcli", "-t", "-f", "DEVICE,TYPE", "device", "status"], timeout=30)
    if result.returncode != 0:
        return None
    for line in command_stdout_lines(result):
        parts = line.split(":")
        if len(parts) >= 2 and parts[1] == "wifi":
            return parts[0]
    return None


def restore_wifi_connection(initial_ssid: str | None, config: dict[str, Any], settle_seconds: int, retries: int) -> None:
    if initial_ssid:
        profiles = [WifiProfile.from_dict(item) for item in config.get("wifi_profiles", [])]
        profile = next((item for item in profiles if item.ssid == initial_ssid), WifiProfile(initial_ssid, "", initial_ssid))
        connect_wifi(profile, settle_seconds, retries)
        logging.info("Koneksi Wi-Fi dikembalikan ke SSID awal: %s", initial_ssid)
    else:
        disconnect_wifi()
        logging.info("Wi-Fi diputus karena sebelum tes tidak ada SSID Wi-Fi aktif.")


def is_connected_to(ssid: str) -> bool:
    system = platform.system().lower()

    if system == "windows":
        result = run(["netsh", "wlan", "show", "interfaces"], timeout=30)
        return result.returncode == 0 and ssid.lower() in (result.stdout or "").lower()

    if system == "darwin":
        device = get_macos_wifi_device()
        result = run(["networksetup", "-getairportnetwork", device], timeout=30)
        return result.returncode == 0 and ssid.lower() in (result.stdout or "").lower()

    if system == "linux":
        result = run(["nmcli", "-t", "-f", "active,ssid", "dev", "wifi"], timeout=30)
        return result.returncode == 0 and f"yes:{ssid}".lower() in (result.stdout or "").lower()

    return False


def perform_speedtest() -> dict[str, float]:
    ensure_standard_streams()
    try:
        import speedtest
    except ImportError as exc:
        raise WifiMonitorError(ERROR_SPEEDTEST_FAILED, "Dependency speedtest-cli belum tersedia.") from exc

    os.environ.update(speedtest_env())
    try:
        tester = speedtest.Speedtest(timeout=60, secure=True)
        tester.get_best_server()
        download = tester.download()
        upload = tester.upload()
        data = tester.results.dict()
    except Exception as exc:
        message = str(exc)
        error_type = ERROR_NO_INTERNET if "urlopen error" in message.lower() else ERROR_SPEEDTEST_FAILED
        raise WifiMonitorError(error_type, message) from exc

    return {
        "download_mbps": round(float(data.get("download", download)) / 1_000_000, 2),
        "upload_mbps": round(float(data.get("upload", upload)) / 1_000_000, 2),
        "ping_ms": round(float(data.get("ping", 0)), 2),
    }


def ensure_standard_streams() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if stream is not None:
            continue
        replacement = open(os.devnull, "w", encoding="utf-8")
        setattr(sys, stream_name, replacement)
        _STANDARD_STREAM_REFS.append(replacement)


def speedtest_env() -> dict[str, str]:
    env = os.environ.copy()
    try:
        import certifi
    except ImportError:
        return env

    ca_bundle = certifi.where()
    env.setdefault("SSL_CERT_FILE", ca_bundle)
    env.setdefault("REQUESTS_CA_BUNDLE", ca_bundle)
    return env


def append_csv(row: dict[str, Any]) -> None:
    with CSV_FILE.open("a", newline="", encoding="utf-8-sig") as f:
        csv.DictWriter(f, fieldnames=HEADERS).writerow({header: row.get(header, "") for header in HEADERS})


def read_rows() -> list[dict[str, str]]:
    with CSV_FILE.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def day_name(date_text: str) -> str:
    names = {
        0: "Senin",
        1: "Selasa",
        2: "Rabu",
        3: "Kamis",
        4: "Jumat",
        5: "Sabtu",
        6: "Minggu",
    }
    try:
        return names[datetime.strptime(date_text, "%Y-%m-%d").weekday()]
    except ValueError:
        return "-"


def safe_float(value: str) -> float | None:
    if value in {"", None}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def sort_report_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return sorted(rows, key=lambda r: (r.get("tanggal", ""), r.get("waktu", ""), r.get("wifi", "")))


def write_excel(rows: list[dict[str, str]], month_key: str) -> Path:
    output = REPORT_DIR / f"laporan_wifi_{month_key}.xlsx"
    selected = sort_report_rows([r for r in rows if r["tanggal"].startswith(month_key)])

    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Speedtest"
    ws.append([
        "Tanggal",
        "Hari",
        "Waktu",
        "Komputer",
        "Wi-Fi",
        "Download (Mbps)",
        "Upload (Mbps)",
        "Ping (ms)",
        "Status",
        "Tipe Error",
        "Keterangan",
    ])

    for row in selected:
        ws.append([
            row["tanggal"],
            day_name(row["tanggal"]),
            row["waktu"],
            row["komputer"],
            row["wifi"],
            safe_float(row["download_mbps"]),
            safe_float(row["upload_mbps"]),
            safe_float(row["ping_ms"]),
            row["status"],
            row.get("error_type", ""),
            row["error"],
        ])

    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for cell in ws["K"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [13, 12, 11, 24, 23, 18, 17, 12, 12, 18, 50]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws["M1"] = "Buka Grafik"
    ws["M1"].hyperlink = "#'Grafik'!A1"
    ws["M1"].font = Font(bold=True, color="FFFFFF")
    ws["M1"].fill = PatternFill("solid", fgColor="4472C4")
    ws["M1"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["M"].width = 16

    write_graph_sheet(wb, selected)
    wb.save(output)
    return output


def write_graph_sheet(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("Grafik")
    ws.freeze_panes = "B3"
    ws["A1"] = "Grafik Kecepatan Wi-Fi"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Baris = tanggal/hari, kolom = jam. Nilai memakai rata-rata jika ada beberapa Wi-Fi pada jam yang sama."

    times = sorted({row["waktu"][:5] for row in rows if row.get("waktu")})
    dates = sorted({row["tanggal"] for row in rows if row.get("tanggal")})
    metrics = [
        ("Download (Mbps)", "download_mbps", True),
        ("Upload (Mbps)", "upload_mbps", True),
        ("Ping (ms)", "ping_ms", False),
    ]

    start_row = 4
    for title, key, higher_is_better in metrics:
        start_row = write_metric_matrix(ws, rows, dates, times, title, key, higher_is_better, start_row)
        start_row += 3
    if dates and times:
        start_row = write_daily_charts(ws, rows, dates, times, start_row + 1)

    ws.column_dimensions["A"].width = 24
    for column in range(2, len(times) + 2):
        ws.column_dimensions[get_column_letter(column)].width = 11


def write_metric_matrix(
    ws: Any,
    rows: list[dict[str, str]],
    dates: list[str],
    times: list[str],
    title: str,
    metric_key: str,
    higher_is_better: bool,
    start_row: int,
) -> int:
    ws.cell(start_row, 1, title).font = Font(bold=True, size=12)
    header_row = start_row + 1
    ws.cell(header_row, 1, "Tanggal / Hari")
    for column, time_label in enumerate(times, start=2):
        ws.cell(header_row, column, time_label)

    buckets: dict[tuple[str, str], list[float]] = {}
    for row in rows:
        value = safe_float(row.get(metric_key, ""))
        if value is None:
            continue
        key = (row["tanggal"], row["waktu"][:5])
        buckets.setdefault(key, []).append(value)

    for row_index, date_text in enumerate(dates, start=header_row + 1):
        ws.cell(row_index, 1, f"{date_text} ({day_name(date_text)})")
        for column, time_label in enumerate(times, start=2):
            values = buckets.get((date_text, time_label), [])
            if values:
                ws.cell(row_index, column, round(sum(values) / len(values), 2))

    end_row = header_row + len(dates)
    end_column = max(2, len(times) + 1)
    for row in ws.iter_rows(min_row=header_row, max_row=end_row, min_col=1, max_col=end_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    if dates and times:
        data_range = f"B{header_row + 1}:{get_column_letter(end_column)}{end_row}"
        start_color, end_color = ("F8696B", "63BE7B")
        if not higher_is_better:
            start_color, end_color = end_color, start_color
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min",
                start_color=start_color,
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color=end_color,
            ),
        )

    return end_row


def write_daily_charts(
    ws: Any,
    rows: list[dict[str, str]],
    dates: list[str],
    times: list[str],
    start_row: int,
) -> int:
    ws.cell(start_row, 1, "Plot Harian").font = Font(bold=True, size=12)
    row_anchor = start_row + 2
    for date_text in dates:
        row_anchor = write_daily_chart_pair(ws, rows, date_text, times, row_anchor) + 3
    return row_anchor


def write_daily_chart_pair(
    ws: Any,
    rows: list[dict[str, str]],
    date_text: str,
    times: list[str],
    start_row: int,
) -> int:
    label = f"{day_name(date_text)}, {date_text}"
    ws.cell(start_row, 1, label).font = Font(bold=True)
    header_row = start_row + 1
    headers = ["Jam", "Download (Mbps)", "Upload (Mbps)", "Ping (ms)"]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, column, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    data_row = header_row + 1
    for time_label in times:
        download = average_metric(rows, date_text, time_label, "download_mbps")
        upload = average_metric(rows, date_text, time_label, "upload_mbps")
        ping = average_metric(rows, date_text, time_label, "ping_ms")
        if download is None and upload is None and ping is None:
            continue
        ws.cell(data_row, 1, time_label)
        ws.cell(data_row, 2, download)
        ws.cell(data_row, 3, upload)
        ws.cell(data_row, 4, ping)
        data_row += 1

    if data_row > header_row + 1:
        categories = Reference(ws, min_col=1, min_row=header_row + 1, max_row=data_row - 1)

        speed_chart = LineChart()
        speed_chart.title = f"Speed - {label}"
        speed_chart.y_axis.title = "Speed (Mbps)"
        speed_chart.x_axis.title = "Jam"
        speed_chart.height = 7
        speed_chart.width = 18
        speed_data = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=data_row - 1)
        speed_chart.add_data(speed_data, titles_from_data=True)
        speed_chart.set_categories(categories)
        style_line_series(speed_chart, ["4472C4", "ED7D31"])
        ws.add_chart(speed_chart, f"F{start_row}")

        ping_chart = LineChart()
        ping_chart.title = f"Ping - {label}"
        ping_chart.y_axis.title = "Ping (ms)"
        ping_chart.x_axis.title = "Jam"
        ping_chart.height = 7
        ping_chart.width = 18
        ping_data = Reference(ws, min_col=4, min_row=header_row, max_row=data_row - 1)
        ping_chart.add_data(ping_data, titles_from_data=True)
        ping_chart.set_categories(categories)
        style_line_series(ping_chart, ["70AD47"])
        ping_chart.legend = None
        ws.add_chart(ping_chart, f"P{start_row}")

    for column, width in enumerate([11, 18, 18, 12], start=1):
        ws.column_dimensions[get_column_letter(column)].width = max(ws.column_dimensions[get_column_letter(column)].width or 0, width)
    return max(data_row, start_row + 16)


def average_metric(rows: list[dict[str, str]], date_text: str, time_label: str, metric_key: str) -> float | None:
    values = [
        value
        for row in rows
        if row.get("tanggal") == date_text and row.get("waktu", "")[:5] == time_label
        for value in [safe_float(row.get(metric_key, ""))]
        if value is not None
    ]
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def style_line_series(chart: LineChart, colors: list[str]) -> None:
    for series, color in zip(chart.series, colors):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 25000
        series.marker.symbol = "circle"
        series.marker.size = 6
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color


def write_daily_pdf(rows: list[dict[str, str]], date_key: str) -> Path:
    output = REPORT_DIR / f"laporan_wifi_{date_key}.pdf"
    selected = sort_report_rows([r for r in rows if r["tanggal"] == date_key])

    doc = SimpleDocTemplate(
        str(output),
        pagesize=landscape(A3),
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    cell_style = styles["BodyText"]
    cell_style.fontSize = 7
    cell_style.leading = 8
    story = [
        Paragraph(f"Laporan Kecepatan Wi-Fi - {date_key}", styles["Title"]),
        Spacer(1, 5 * mm),
    ]

    table_data = [[
        "Tanggal", "Hari", "Waktu", "Komputer", "Wi-Fi",
        "Download", "Upload", "Ping", "Status", "Tipe Error", "Keterangan"
    ]]
    for row in selected:
        table_data.append([
            row["tanggal"],
            day_name(row["tanggal"]),
            row["waktu"],
            Paragraph(row["komputer"], cell_style),
            Paragraph(row["wifi"], cell_style),
            f'{row["download_mbps"]} Mbps' if row["download_mbps"] else "-",
            f'{row["upload_mbps"]} Mbps' if row["upload_mbps"] else "-",
            f'{row["ping_ms"]} ms' if row["ping_ms"] else "-",
            row["status"],
            row.get("error_type", "-") or "-",
            Paragraph(row["error"] or "-", cell_style),
        ])

    table = LongTable(
        table_data,
        repeatRows=1,
        colWidths=[
            25 * mm, 22 * mm, 18 * mm, 42 * mm, 40 * mm,
            27 * mm, 27 * mm, 20 * mm, 18 * mm, 28 * mm, 103 * mm,
        ],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
    ]))
    story.append(table)
    doc.build(story)
    return output


def shutdown_computer(delay_seconds: int = 30) -> None:
    system = platform.system().lower()
    logging.info("Memulai shutdown otomatis untuk %s.", system)

    if system == "windows":
        subprocess.Popen([
            "shutdown",
            "/s",
            "/t",
            str(delay_seconds),
            "/c",
            "Pengujian Wi-Fi selesai dan laporan berhasil disimpan.",
        ], creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
    elif system in {"darwin", "linux"}:
        subprocess.Popen(["shutdown", "-h", f"+{max(1, delay_seconds // 60)}"])
    else:
        raise RuntimeError(f"Shutdown belum didukung untuk OS: {platform.system()}")


def maybe_send_email_report(
    config: dict[str, Any],
    report_date: date,
    attachments: list[Path],
    result_callback: Any | None = None,
) -> bool:
    email_config = config.get("email_report") or {}
    if not bool(email_config.get("enabled", False)):
        return True

    if not bool(email_config.get("send_after_final", True)):
        logging.info("Email laporan aktif, tetapi send_after_final nonaktif.")
        return True

    if not email_report_due(email_config, report_date):
        logging.info("Email laporan dilewati karena tanggal %s tidak sesuai jadwal kirim.", report_date.isoformat())
        return True

    try:
        send_email_report(email_config, report_date, attachments)
        message = f"Email laporan berhasil dikirim ke {', '.join(parse_recipients(str(email_config.get('to', ''))))}."
        logging.info(message)
        if result_callback is not None:
            result_callback({"status": "INFO", "wifi": "Email", "error": message})
        return True
    except Exception:
        logging.exception("Gagal mengirim email laporan.")
        if result_callback is not None:
            result_callback({"status": "INFO", "wifi": "Email", "error": "Gagal mengirim email laporan. Cek logs/monitor.log."})
        return False


def email_report_due(email_config: dict[str, Any], report_date: date) -> bool:
    weekdays = {int(day) for day in email_config.get("weekdays", []) if str(day).strip() != ""}
    dates = {str(item).strip() for item in email_config.get("dates", []) if str(item).strip()}

    if report_date.isoformat() in dates:
        return True
    if report_date.weekday() in weekdays:
        return True
    return not weekdays and not dates


def send_email_report(email_config: dict[str, Any], report_date: date, attachments: list[Path]) -> None:
    host = str(email_config.get("smtp_host", "")).strip()
    port = int(email_config.get("smtp_port", 587))
    username = str(email_config.get("username", "")).strip()
    password = str(email_config.get("password", ""))
    sender = str(email_config.get("from", "")).strip() or username
    recipients = parse_recipients(str(email_config.get("to", "")))
    subject_template = str(email_config.get("subject", "Laporan Wi-Fi {date}"))

    if not host:
        raise WifiMonitorError(ERROR_EMAIL_FAILED, "SMTP host belum diisi.")
    if not sender:
        raise WifiMonitorError(ERROR_EMAIL_FAILED, "Email pengirim belum diisi.")
    if not recipients:
        raise WifiMonitorError(ERROR_EMAIL_FAILED, "Email tujuan belum diisi.")

    selected = [path for path in attachments if path.exists()]
    if not selected:
        raise WifiMonitorError(ERROR_EMAIL_FAILED, "Tidak ada file laporan yang bisa dilampirkan.")

    message = EmailMessage()
    message["From"] = sender
    message["To"] = ", ".join(recipients)
    message["Subject"] = format_email_subject(subject_template, report_date)
    message.set_content(
        "Terlampir laporan Wi-Fi Speed Monitor.\n\n"
        f"Tanggal laporan: {report_date.isoformat()}\n"
    )

    for path in selected:
        content_type, _encoding = mimetypes.guess_type(path.name)
        maintype, subtype = (content_type or "application/octet-stream").split("/", 1)
        message.add_attachment(path.read_bytes(), maintype=maintype, subtype=subtype, filename=path.name)

    use_ssl = bool(email_config.get("use_ssl", False))
    use_tls = bool(email_config.get("use_tls", True))

    if use_ssl:
        with smtplib.SMTP_SSL(host, port, timeout=60) as smtp:
            login_smtp(smtp, username, password)
            smtp.send_message(message)
    else:
        with smtplib.SMTP(host, port, timeout=60) as smtp:
            smtp.ehlo()
            if use_tls:
                smtp.starttls()
                smtp.ehlo()
            login_smtp(smtp, username, password)
            smtp.send_message(message)


def login_smtp(smtp: smtplib.SMTP, username: str, password: str) -> None:
    if username or password:
        smtp.login(username, password)


def parse_recipients(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;,]", value) if item.strip()]


def format_email_subject(template: str, report_date: date) -> str:
    try:
        return template.format(date=report_date.isoformat())
    except Exception:
        logging.warning("Format subject email tidak valid: %s", template)
        return f"Laporan Wi-Fi {report_date.isoformat()}"


def selected_email_attachments(config: dict[str, Any], excel_report: Path, pdf_report: Path) -> list[Path]:
    email_config = config.get("email_report") or {}
    attachments = []
    if bool(email_config.get("attach_excel", True)):
        attachments.append(excel_report)
    if bool(email_config.get("attach_pdf", True)):
        attachments.append(pdf_report)
    return attachments


def should_skip_for_schedule_lifecycle(config: dict[str, Any], result_callback: Any | None = None) -> bool:
    schedule = config.get("schedule") or {}
    if not schedule:
        return False

    today = date.today()
    active_start = parse_schedule_date(str(schedule.get("active_start_date", "")).strip())
    active_until = parse_schedule_date(str(schedule.get("active_until_date", "")).strip())

    active_days = int(schedule.get("active_days", 0) or 0)
    if active_until is None and active_days > 0:
        start = active_start or today
        active_until = start + timedelta(days=active_days - 1)

    if active_start and today < active_start:
        message = f"Jadwal belum aktif. Mulai aktif pada {active_start.isoformat()}."
        logging.info(message)
        if result_callback is not None:
            result_callback({"status": "INFO", "wifi": "Schedule", "error": message})
        return True

    if active_until and today > active_until:
        message = f"Masa aktif jadwal selesai pada {active_until.isoformat()}; menghapus jadwal OS."
        logging.info(message)
        if result_callback is not None:
            result_callback({"status": "INFO", "wifi": "Schedule", "error": message})
        try:
            from install_schedule import uninstall_for_current_os

            uninstall_for_current_os()
            logging.info("Jadwal OS berhasil dihapus otomatis.")
        except Exception:
            logging.exception("Gagal menghapus jadwal OS otomatis.")
        return True

    return False


def parse_schedule_date(value: str) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        logging.warning("Tanggal jadwal tidak valid: %s", value)
        return None


def run_monitor(
    config: dict[str, Any],
    final: bool = False,
    result_callback: Any | None = None,
    enforce_schedule_lifecycle: bool = True,
) -> int:
    if enforce_schedule_lifecycle and should_skip_for_schedule_lifecycle(config, result_callback):
        return 0

    computer_name = config.get("computer_name") or platform.node() or "Komputer"
    settle_seconds = int(config.get("settle_seconds", 20))
    gap_seconds = int(config.get("gap_between_tests_seconds", 20))
    connection_retries = int(config.get("connection_retries", 2))
    restore_after_tests = bool(config.get("restore_connection_after_tests", True))
    initial_ssid = current_wifi_ssid() if restore_after_tests else None
    if restore_after_tests:
        logging.info("SSID awal sebelum tes: %s", initial_ssid or "tidak ada")
    all_tests_ok = True

    test_items: list[tuple[str, WifiProfile | None, str]] = []
    if bool(config.get("test_current_connection", False)):
        label = str(config.get("current_connection_label") or "Ethernet / Koneksi Aktif")
        test_items.append(("current", None, label))
    for item in config["wifi_profiles"]:
        profile = WifiProfile.from_dict(item)
        test_items.append(("wifi", profile, profile.label or profile.ssid))

    wifi_tests_run = any(kind == "wifi" for kind, _profile, _label in test_items)

    for index, (kind, profile, label) in enumerate(test_items):
        now = datetime.now()
        row: dict[str, Any] = {
            "tanggal": now.strftime("%Y-%m-%d"),
            "waktu": now.strftime("%H:%M:%S"),
            "komputer": computer_name,
            "wifi": label,
            "download_mbps": "",
            "upload_mbps": "",
            "ping_ms": "",
            "status": "GAGAL",
            "error_type": "",
            "error": "",
        }

        try:
            if kind == "wifi" and profile is not None:
                connect_wifi(profile, settle_seconds, connection_retries)
            else:
                logging.info("Menguji koneksi aktif tanpa mengganti Wi-Fi.")
            result = perform_speedtest()
            row.update(result)
            row["status"] = "OK"
            logging.info(
                "%s / %s: download=%s upload=%s",
                computer_name,
                label,
                row["download_mbps"],
                row["upload_mbps"],
            )
        except Exception as exc:
            all_tests_ok = False
            row["error_type"] = getattr(exc, "error_type", ERROR_UNKNOWN)
            row["error"] = str(exc)
            logging.exception("Pengujian gagal untuk %s", label)

        append_csv(row)
        if result_callback is not None:
            result_callback(dict(row))

        if index < len(test_items) - 1:
            time.sleep(gap_seconds)

    if restore_after_tests and wifi_tests_run:
        try:
            restore_wifi_connection(initial_ssid, config, settle_seconds, connection_retries)
            if result_callback is not None:
                result_callback({"status": "INFO", "wifi": "Restore", "error": f"Koneksi dikembalikan ke {initial_ssid or 'mode non-Wi-Fi'}."})
        except Exception:
            logging.exception("Gagal mengembalikan koneksi awal.")

    reports_ok = False
    email_ok = True
    try:
        rows = read_rows()
        now = datetime.now()
        excel_report = write_excel(rows, now.strftime("%Y-%m"))
        pdf_report = write_daily_pdf(rows, now.strftime("%Y-%m-%d"))
        reports_ok = True
        if final:
            attachments = selected_email_attachments(config, excel_report, pdf_report)
            email_ok = maybe_send_email_report(config, now.date(), attachments, result_callback)
    except Exception:
        logging.exception("Gagal membuat laporan Excel/PDF.")

    if final and bool(config.get("shutdown_after_final", False)):
        if all_tests_ok and reports_ok and email_ok:
            shutdown_computer(int(config.get("shutdown_delay_seconds", 30)))
        else:
            logging.error("Shutdown dibatalkan karena tes, laporan, atau email gagal.")

    return 0 if all_tests_ok and reports_ok and email_ok else 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Monitor speedtest Wi-Fi lintas OS.")
    parser.add_argument("--config", default=str(CONFIG_FILE), help="Lokasi file konfigurasi JSON.")
    parser.add_argument("--final", action="store_true", help="Run final; shutdown jika diaktifkan di config.")
    args = parser.parse_args()

    setup()
    config = load_config(Path(args.config))
    return run_monitor(config, final=args.final)


if __name__ == "__main__":
    raise SystemExit(main())
