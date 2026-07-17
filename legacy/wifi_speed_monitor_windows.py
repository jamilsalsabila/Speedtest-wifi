from __future__ import annotations

import argparse
import csv
import json
import logging
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = Path(__file__).resolve().parent
CONFIG_FILE = BASE_DIR / "config.json"
DATA_DIR = BASE_DIR / "data"
REPORT_DIR = BASE_DIR / "reports"
LOG_DIR = BASE_DIR / "logs"
CSV_FILE = DATA_DIR / "speedtest_log.csv"

HEADERS = [
    "tanggal", "waktu", "komputer", "wifi",
    "download_mbps", "upload_mbps", "ping_ms",
    "status", "error"
]


def setup() -> None:
    for folder in (DATA_DIR, REPORT_DIR, LOG_DIR):
        folder.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        filename=LOG_DIR / "monitor.log",
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        encoding="utf-8",
    )

    if not CSV_FILE.exists():
        with CSV_FILE.open("w", newline="", encoding="utf-8-sig") as f:
            csv.DictWriter(f, fieldnames=HEADERS).writeheader()


def run(args: list[str], timeout: int = 240) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        args,
        capture_output=True,
        text=True,
        timeout=timeout,
        check=False,
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )


def load_config() -> dict[str, Any]:
    if not CONFIG_FILE.exists():
        raise FileNotFoundError("config.json belum dibuat.")
    with CONFIG_FILE.open(encoding="utf-8") as f:
        config = json.load(f)

    if not config.get("wifi_profiles"):
        raise ValueError("wifi_profiles di config.json masih kosong.")
    return config


def connect_wifi(ssid: str, settle_seconds: int) -> None:
    logging.info("Menghubungkan ke SSID %s", ssid)

    result = run(["netsh", "wlan", "connect", f"name={ssid}", f"ssid={ssid}"], timeout=45)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip())

    time.sleep(settle_seconds)

    status = run(["netsh", "wlan", "show", "interfaces"], timeout=30)
    output = status.stdout.lower()

    if "state" not in output and "status" not in output:
        raise RuntimeError("Tidak dapat membaca status Wi-Fi Windows.")

    if ssid.lower() not in output:
        raise RuntimeError(f"Komputer belum tersambung ke {ssid}.")


def perform_speedtest() -> dict[str, float]:
    # Menggunakan paket Python speedtest-cli.
    result = run([sys.executable, "-m", "speedtest", "--json", "--secure"], timeout=240)

    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip())

    data = json.loads(result.stdout)
    return {
        "download_mbps": round(float(data["download"]) / 1_000_000, 2),
        "upload_mbps": round(float(data["upload"]) / 1_000_000, 2),
        "ping_ms": round(float(data.get("ping", 0)), 2),
    }


def append_csv(row: dict[str, Any]) -> None:
    with CSV_FILE.open("a", newline="", encoding="utf-8-sig") as f:
        csv.DictWriter(f, fieldnames=HEADERS).writerow(row)


def read_rows() -> list[dict[str, str]]:
    with CSV_FILE.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_excel(rows: list[dict[str, str]], month_key: str) -> Path:
    output = REPORT_DIR / f"laporan_wifi_{month_key}.xlsx"
    selected = [r for r in rows if r["tanggal"].startswith(month_key)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Speedtest"

    display_headers = [
        "Tanggal", "Waktu", "Komputer", "Wi-Fi",
        "Download (Mbps)", "Upload (Mbps)", "Ping (ms)",
        "Status", "Keterangan"
    ]
    ws.append(display_headers)

    for row in selected:
        ws.append([
            row["tanggal"],
            row["waktu"],
            row["komputer"],
            row["wifi"],
            float(row["download_mbps"]) if row["download_mbps"] else None,
            float(row["upload_mbps"]) if row["upload_mbps"] else None,
            float(row["ping_ms"]) if row["ping_ms"] else None,
            row["status"],
            row["error"],
        ])

    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    widths = [13, 11, 24, 23, 18, 17, 12, 12, 50]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output)
    return output


def write_daily_pdf(rows: list[dict[str, str]], date_key: str) -> Path:
    output = REPORT_DIR / f"laporan_wifi_{date_key}.pdf"
    selected = [r for r in rows if r["tanggal"] == date_key]

    doc = SimpleDocTemplate(
        str(output),
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Laporan Kecepatan Wi-Fi — {date_key}", styles["Title"]),
        Spacer(1, 5 * mm),
    ]

    table_data = [[
        "Tanggal", "Waktu", "Komputer", "Wi-Fi",
        "Download", "Upload", "Ping", "Status"
    ]]

    for row in selected:
        table_data.append([
            row["tanggal"],
            row["waktu"],
            row["komputer"],
            row["wifi"],
            f'{row["download_mbps"]} Mbps' if row["download_mbps"] else "-",
            f'{row["upload_mbps"]} Mbps' if row["upload_mbps"] else "-",
            f'{row["ping_ms"]} ms' if row["ping_ms"] else "-",
            row["status"],
        ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[25*mm, 20*mm, 42*mm, 40*mm, 31*mm, 29*mm, 23*mm, 22*mm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
    ]))
    story.append(table)
    doc.build(story)
    return output


def shutdown_windows(delay_seconds: int = 30) -> None:
    logging.info("Memulai shutdown otomatis.")
    subprocess.Popen(
        ["shutdown", "/s", "/t", str(delay_seconds), "/c",
         "Pengujian Wi-Fi pukul 21.00 selesai dan laporan berhasil disimpan."],
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--final",
        action="store_true",
        help="Pengujian terakhir pukul 21.00; shutdown jika seluruh proses berhasil.",
    )
    args = parser.parse_args()

    setup()
    config = load_config()

    computer_name = config.get("computer_name", "Komputer")
    settle_seconds = int(config.get("settle_seconds", 20))
    gap_seconds = int(config.get("gap_between_tests_seconds", 20))
    all_tests_ok = True

    for index, item in enumerate(config["wifi_profiles"]):
        ssid = item["ssid"]
        label = item.get("label", ssid)
        now = datetime.now()

        row: dict[str, Any] = {
            "tanggal": now.strftime("%Y-%m-%d"),
            "waktu": now.strftime("%H:%M:%S"),
            "komputer": computer_name,
            "wifi": label,
            "download_mbps": "",
            "upload_mbps": "",
            "ping_ms": "",
            "status": "GAGAL",
            "error": "",
        }

        try:
            connect_wifi(ssid, settle_seconds)
            result = perform_speedtest()
            row.update(result)
            row["status"] = "OK"
            logging.info(
                "%s / %s: download=%s upload=%s",
                computer_name, label,
                row["download_mbps"], row["upload_mbps"],
            )
        except Exception as exc:
            all_tests_ok = False
            row["error"] = str(exc)
            logging.exception("Pengujian gagal untuk %s", label)

        append_csv(row)

        if index < len(config["wifi_profiles"]) - 1:
            time.sleep(gap_seconds)

    reports_ok = False
    try:
        rows = read_rows()
        now = datetime.now()
        write_excel(rows, now.strftime("%Y-%m"))
        write_daily_pdf(rows, now.strftime("%Y-%m-%d"))
        reports_ok = True
    except Exception:
        logging.exception("Gagal membuat laporan Excel/PDF.")

    if args.final:
        if all_tests_ok and reports_ok:
            shutdown_windows(int(config.get("shutdown_delay_seconds", 30)))
        else:
            logging.error(
                "Shutdown dibatalkan karena tes atau penyimpanan laporan gagal."
            )

    return 0 if all_tests_ok and reports_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
